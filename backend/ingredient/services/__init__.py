from io import BytesIO
from typing import Any
from uuid import UUID

from django.core.files.uploadedfile import UploadedFile
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from attachment.models import Attachment
from attachment.services import AttachmentService
from utils.permissions.decorators import sync_user_feature
from dish.schemas.requests import DishIngredientSuggestionSchema
from dish.orm.dish import DishORM
from exceptions.ingredient import IngredientImportFileInvalid, IngredientImportFileRequired, IngredientIsNotPending
from exceptions.ingredient import (
    IngredientDoesNotExist,
    IngredientIsNotDeleted,
    IngredientIsReferenced,
    IngredientSuggestionNotFound,
    IngredientNameAlreadyExists,
)
from exceptions.dishes import DishIngredientNotFoundException
from ingredient.constants import (
    HIGH_FAT_KEYWORDS,
    HIGH_PROTEIN_KEYWORDS,
    HIGH_SUGAR_KEYWORDS,
    IMPORT_COLUMNS,
    DISH_NUTRIENT_FIELDS
)
from ingredient.models import Ingredient, IngredientImportStatusEnum, IngredientSuggestion
from ingredient.orm.ingredient import IngredientORM
from ingredient.schemas.requests import (
    FilterIngredientSchema,
    IngredientAliasCreateSchema,
    IngredientSuggestionApproveAliasSchema,
    IngredientSuggestionApproveNewSchema,
    IngredientSuggestionOrderBySchema,
    IngredientSchema,
    IngredientSuggestionFilterSchema,
    IngredientSuggestionResolveSchema,
    OrderByIngredientSchema,
    UserIngredientPreferenceSchema,
)
from ingredient.schemas.responses import (
    IngredientAliasResponse,
    IngredientAutocompleteItem,
    IngredientImportError,
    IngredientImportResult,
    IngredientSearchItem,
    IngredientSuggestionCandidateResponse,
    IngredientSuggestionResponse,
    UserIngredientPreferenceResponse,
)
from utils.enums import IngredientCategoryEnum,IngredientSourceEnum
from utils.functions.remove_accents import remove_accents
from utils.types import TUser
from django.db import transaction
from difflib import SequenceMatcher


class IngredientQueryService:
    def __init__(self):
        self.orm = IngredientORM()

    def get_all(self, filter: FilterIngredientSchema, order_by: OrderByIngredientSchema):
        return self.orm.get_all_ingredients(filter=filter, order_by=order_by)

    def get_all_for_chef(self, user: TUser, filter: FilterIngredientSchema, order_by: OrderByIngredientSchema):
        # lấy cả ingredient thường + pending suggestion của chính chef đó
   
        # =========================
        # 1. USDA
        # =========================
        qs = Ingredient.objects.filter(
            deleted=False,
            category__isnull=False,
        )

        if filter:
            qs = qs.filter(filter.get_filter_expression())

        if order_by.order_by == "energy":
            qs = qs.order_by("-energy")
        else:
            qs = qs.order_by("-updated_at")

        usda_items = list(qs[:100])  # tránh query quá lớn

        # =========================
        # 2. USER SUGGESTION
        # =========================
        suggestion_qs = IngredientSuggestion.objects.filter(
            created_by=user,
            status=IngredientImportStatusEnum.PENDING,
            deleted=False,
        )
        #debug print suggestion
        print(f"Found {suggestion_qs.count()} pending suggestions for user {user.id}")

        if filter.categories:
            categories = [c.strip() for c in filter.categories.split(",") if c.strip()]
            suggestion_qs = suggestion_qs.filter(suggested_category__in=categories)

        suggestions = list(suggestion_qs[:100])

        # =========================
        # 3. BUILD RESULT
        # =========================
        results = []

        # USDA
        for i in usda_items:
            results.append({
                "ingredient_uid": i.uid,
                "suggestion_uid": None,
                "name": i.name,
                "category": i.category,
                "matched_alias": None,
                "score": 1.0,
                "source": "USDA",
                "approval_status": "APPROVED",
                "is_suggestion": False,
            })

        # SUGGESTION
        for s in suggestions:
            if not s.suggested_category:
                continue  # 🔥 tránh crash enum

            results.append({
                "ingredient_uid": None,
                "suggestion_uid": s.uid,
                "name": s.suggested_name,
                "category": s.suggested_category,
                "matched_alias": None,
                "score": 0.7,  # thấp hơn USDA
                "source": "CHEF_SUGGESTION",
                "approval_status": s.status,
                "is_suggestion": True,
            })

        # =========================
        # 4. DEDUP (QUAN TRỌNG)
        # =========================
        seen = set()
        deduped = []

        for item in results:
            key = item["name"].strip().lower()

            if key in seen:
                continue

            seen.add(key)
            deduped.append(item)

        # =========================
        # 5. SORT FINAL
        # =========================
        def sort_key(item):
            source_priority = 0 if item["source"] == "USDA" else 1

            if order_by.order_by == "energy":
                energy = getattr(item, "energy", 0) if hasattr(item, "energy") else 0
                return (source_priority, -energy)
            else:
                return (source_priority, item["name"].lower())

        deduped = sorted(deduped, key=sort_key)

        return [IngredientSearchItem(**item) for item in deduped]
    def get_by_uid(self, uid: UUID):
        ingredient = self.orm.get_ingredient_by_uid(uid=uid)
        if not ingredient:
            raise IngredientDoesNotExist
        return ingredient
    
    def compute_suggestion_score(self, query: str, name: str):
        base = SequenceMatcher(None, query.lower(), name.lower()).ratio()

        # map về range thấp hơn USDA
        return 0.55 + base * 0.25

    def search(
        self,
        user: TUser,
        query: str,
        filter: FilterIngredientSchema,
        limit: int = 10,
    ):
        
            # 🔥 GUARD CLAUSE
        query = (query or "").strip()
        if not query:
            return []
        fetch_limit = limit * 2 if limit > 0 else 1000
        # Extract category from filter if present
        category = getattr(filter, "categories", None)
        #. 1 USDA
        # Use find_suggestion_candidates for better relevance-based search
        usda_results = self.orm.find_suggestion_candidates(
            suggested_name=query,
            limit=fetch_limit,
            category=category,
        )
    
        # =========================
        # 2. CHEF SUGGESTION (ONLY OWN)
        # =========================
        suggestions = self.orm.find_user_pending_suggestions(
            user=user,
            query=query,
            limit=fetch_limit,
            category=category,
        )

        # =========================
        # 3. MERGE
        # =========================
        results = []

        # USDA
        for item in usda_results:
            results.append({
                **item,
                "ingredient_uid": item["uid"],
                "suggestion_uid": None,
                "source": "USDA",
                "approval_status": "APPROVED",
                "is_suggestion": False,
            })

        # SUGGESTION
        for s in suggestions:
            score = self.compute_suggestion_score(query, s.suggested_name)

            results.append({
                "ingredient_uid": None,
                "suggestion_uid": s.uid,
                "name": s.suggested_name,
                "category": s.suggested_category if s.suggested_category else "UNKNOWN",
                "matched_alias": None,
                "score": round(score, 4),
                "source": "CHEF_SUGGESTION",
                "approval_status": s.status,
                "is_suggestion": True,
            })

        seen = set()
        deduped = []

        for item in results:
            key = (item["name"].strip().lower(), item["source"])

            if key in seen:
                continue

            seen.add(key)
            deduped.append(item)

        results = deduped

        return [IngredientSearchItem(**item) for item in results[:limit]]

    def autocomplete(self, query: str, limit: int = 10):
        results = self.orm.autocomplete_ingredients(query=query, limit=limit)
        return [IngredientAutocompleteItem(**item) for item in results]

    def list_aliases(self, search: str | None = None):
        return [
            IngredientAliasResponse(
                uid=item.uid,
                ingredient_uid=item.ingredient.uid,
                ingredient_name=item.ingredient.name,
                alias=item.alias,
            )
            for item in self.orm.list_ingredient_aliases(search=search)
        ]


class IngredientCommandService:
    def __init__(self, query_service: IngredientQueryService):
        self.orm = IngredientORM()
        self.query_service = query_service

    def create_ingredient(self, user: TUser, payload: IngredientSchema):
        name = payload.name.strip()
        if Ingredient.objects.filter(name__iexact=name, deleted=False).exists():
            raise IngredientNameAlreadyExists

        return self.orm.create_new_ingredient(user=user, payload=payload)

    def update_ingredient(self, user: TUser, uid: UUID, payload: IngredientSchema):
        return self.orm.update_ingredient(
            user=user,
            ingredient=self.query_service.get_by_uid(uid=uid),
            payload=payload,
        )

    def soft_delete_ingredient(self, user: TUser, uid: UUID):
        if self.orm.soft_delete_ingredient(user=user, ingredient=self.query_service.get_by_uid(uid=uid)):
            return True
        raise IngredientIsReferenced

    def hard_delete_ingredient(self, uid: UUID):
        if self.orm.delete_ingredient(ingredient=self.query_service.get_by_uid(uid=uid)):
            return True
        raise IngredientIsReferenced

    def restore_ingredient(self, user: TUser, uid: UUID):
        if self.orm.restore_ingredient(user=user, ingredient=self.query_service.get_by_uid(uid=uid)):
            return True
        raise IngredientIsNotDeleted

    def create_alias(self, user: TUser, payload: IngredientAliasCreateSchema):
        ingredient = self.query_service.get_by_uid(payload.ingredient_uid)
        alias_text = str(getattr(payload, "alias", "")).strip()
        if not alias_text:
            raise ValueError("alias is required")
        alias = self.orm.create_ingredient_alias(ingredient=ingredient, alias=alias_text, user=user)
        return IngredientAliasResponse(
            uid=alias.uid,
            ingredient_uid=alias.ingredient.uid,
            ingredient_name=alias.ingredient.name,
            alias=alias.alias,
        )


class IngredientSuggestionService:
    def __init__(self):
        self.orm = IngredientORM()
        self.dish_orm = DishORM()
        self.attachment_service = AttachmentService()

    def create_suggestion(self, user: TUser, payload: DishIngredientSuggestionSchema, attachment: Attachment | None = None):
        suggestion = IngredientSuggestion.objects.filter(
            suggested_name__iexact=payload.custom_name,
            status=IngredientImportStatusEnum.PENDING,
            created_by=user,
        ).first()
        
        if not suggestion:
            suggestion = self.orm.create_ingredient_suggestion(
                user=user,
                custom_name=payload.custom_name,
                suggested_category=payload.category,
                attachment=attachment,
            )
        return suggestion

    @staticmethod
    def _to_suggestion_response(item: IngredientSuggestion) -> IngredientSuggestionResponse:
        return IngredientSuggestionResponse(
            uid=item.uid,
            name=item.suggested_name,
            category=item.suggested_category,
            status=item.status,
            created_by_id=item.created_by_id,
            verified_by_id=item.verified_by_id,
            verified_at=item.verified_at.isoformat() if item.verified_at else None,
            resolution_note=item.resolution_note,
        )

    def list_suggestions(
        self,
        status: str | None = None,
        search: str | None = None,
        category: IngredientCategoryEnum | None = None,
        order_by: IngredientSuggestionOrderBySchema | None = None,
        created_by_id: int | None = None,
    ):
        items = self.orm.list_ingredient_suggestions(
            status=status,
            search=search,
            category=category,
            created_by_id=created_by_id,
            order_by=order_by,
        )
        return [self._to_suggestion_response(item) for item in items]
        # responses: list[IngredientSuggestionResponse] = []
        # for item in items:
        #     tags = IngredientSuggestionHelper.infer_tags(item.suggested_name)
        #     confidence = IngredientSuggestionHelper.compute_confidence(False, tags)
        #     responses.append(self._build_response(item, [], tags, confidence))
        # return responses

    def soft_delete_ingredient_suggestion(self, user: TUser, uid: UUID):
        suggestion = self.orm.get_ingredient_suggestion_by_uid_and_owner(uid=uid, user=user)
        if not suggestion:
            raise IngredientSuggestionNotFound
        if suggestion.status != IngredientImportStatusEnum.PENDING:
            raise IngredientIsNotPending
        return self.orm.soft_delete_ingredient_suggestion(suggestion=suggestion)

    @transaction.atomic
    def approve_new(self, user: TUser, uid: UUID, payload: IngredientSuggestionApproveNewSchema):
        suggestion = self.orm.get_ingredient_suggestion_by_uid(uid=uid)
        if not suggestion:
            raise IngredientSuggestionNotFound
        dish_ingredient = self.dish_orm.get_dish_ingredient_by_suggestion_uid(suggestion.uid)
        if not dish_ingredient:
            raise DishIngredientNotFoundException
        # handle attachment nếu có
        attachment = None
        if suggestion.attachment:
            attachment = self.attachment_service.handle_attachment(uid=suggestion.attachment.uid)

        # gọi orm ingredient mới
        new_ingredient = self.orm.create_new_ingredient_from_dishingredient(
            user=user,
            category=suggestion.suggested_category,
            dish_ingredient=dish_ingredient,
            attachment=attachment,
        )
        suggestion = self.orm.approve_suggestion_as_new_ingredient(
            suggestion=suggestion,
            user=user,
            resolution_note=payload.resolution_note,
            ingredient=new_ingredient,
        )
        self.dish_orm.update_dish_ingredients_by_suggestion(
            suggestion=suggestion,
            ingredient=new_ingredient,
            status=IngredientImportStatusEnum.APPROVED,
            source=IngredientSourceEnum.CHEF_SUGGESTION,
            user=user
        )

        return {
            "suggestion_uid": suggestion.uid,
            "name": suggestion.suggested_name,
            "category": suggestion.suggested_category,
            "status": suggestion.status,
            "created_by_id": suggestion.created_by_id,
            "verified_by_id": suggestion.verified_by_id,
            "verified_at": suggestion.verified_at.isoformat() if suggestion.verified_at else None,
            "resolution_note": suggestion.resolution_note,
            "new_ingredient_uid": new_ingredient.uid,
        }

    def _scale_usda(self, ingredient, weight):
        if not ingredient.weight or ingredient.weight <= 0:
            return {f: 0.0 for f in DISH_NUTRIENT_FIELDS}

        factor = weight / ingredient.weight

        return {
            field: (getattr(ingredient, field, 0.0) or 0.0) * factor
            for field in DISH_NUTRIENT_FIELDS
        }

    @transaction.atomic
    def approve_alias(self, user: TUser, uid: UUID, payload: IngredientSuggestionApproveAliasSchema):
        suggestion = self.orm.get_ingredient_suggestion_by_uid(uid=uid)
        if not suggestion:
            raise IngredientSuggestionNotFound
        
        ingredient = self.orm.get_ingredient_by_uid(payload.ingredient_uid)
        if not ingredient:
            raise IngredientDoesNotExist
        
        self.orm.approve_suggestion_as_alias(
            suggestion=suggestion,
            user=user,
            ingredient=ingredient,
            resolution_note=payload.resolution_note,
        )

        # 2. lấy tất cả dish ingredient bị ảnh hưởng
        dish_ingredients = self.orm.get_dish_ingredients_by_suggestion(suggestion=suggestion)

        for di in dish_ingredients:
            nutrient_values = self._scale_usda(
                ingredient=ingredient,
                weight=di.weight
            )

            self.dish_orm.apply_alias_resolution_to_dish_ingredient(
                dish_ingredient=di,
                ingredient=ingredient,
                nutrient_values=nutrient_values,
                user=user,
                confidence=1.0,
                approval_status=IngredientImportStatusEnum.APPROVED,
                source=IngredientSourceEnum.USDA,
            )
        return {
            "suggestion_uid": suggestion.uid,
            "name": suggestion.suggested_name,
            "category": suggestion.suggested_category,
            "status": suggestion.status,
            "created_by_id": suggestion.created_by_id,
            "verified_by_id": suggestion.verified_by_id,
            "verified_at": suggestion.verified_at.isoformat() if suggestion.verified_at else None,
            "resolution_note": suggestion.resolution_note,
            "resolved_ingredient_uid": ingredient.uid,
            "resolved_alias_uid": suggestion.resolved_alias.uid
        }


    def reject(self, user: TUser, uid: UUID, rejection_reason: str):
        suggestion = self.orm.get_ingredient_suggestion_by_uid(uid=uid)
        if not suggestion:
            raise IngredientSuggestionNotFound

        self.orm.reject_ingredient_suggestion(
            suggestion=suggestion,
            user=user,
            rejection_reason=rejection_reason,
        )
        # tính lại confidence
        #import dish service để tính lại confidence cho dish ingredient liên quan
        dish_ingredients = self.orm.get_dish_ingredients_by_suggestion(suggestion=suggestion)
        from dish.services import DishService
        dish_service = DishService()

        for di in dish_ingredients:
            values, warnings, confidence = dish_service._process_ingredient_pipeline(
                payload=di,                  
                ingredient=None,
                approval_status=IngredientImportStatusEnum.REJECTED,
                source=IngredientSourceEnum.CHEF_SUGGESTION,
            )
            self.dish_orm.update_dish_ingredients_by_suggestion(
                suggestion=suggestion,  
                ingredient=None,
                status=IngredientImportStatusEnum.REJECTED,
                source=IngredientSourceEnum.CHEF_SUGGESTION,
                user=user,
                confidence=confidence
            )
        return {
            "suggestion_uid": suggestion.uid,
            "name": suggestion.suggested_name,
            "category": suggestion.suggested_category,
            "status": suggestion.status,
            "created_by_id": suggestion.created_by_id,
            "verified_by_id": suggestion.verified_by_id,
            "verified_at": suggestion.verified_at.isoformat() if suggestion.verified_at else None,
            "rejection_reason": rejection_reason
        }


    # def _build_response(
    #     self,
    #     suggestion,
    #     candidates,
    #     tags: dict[str, bool],
    #     confidence: float,
    # ):
    #     return IngredientSuggestionResponse(
    #         uid=suggestion.uid,
    #         name=suggestion.suggested_name,
    #         category=suggestion.suggested_category,
    #         status=suggestion.status,
    #         created_by_id=suggestion.created_by_id,
    #         verified_by_id=suggestion.verified_by_id,
    #         verified_at=suggestion.verified_at.isoformat() if suggestion.verified_at else None,
    #         rejection_reason=suggestion.rejection_reason,
    #         resolved_ingredient_uid=suggestion.ingredient.uid if suggestion.ingredient else None,
    #         resolved_alias_uid=suggestion.resolved_alias.uid if suggestion.resolved_alias else None,
    #         resolution_note=suggestion.resolution_note,
    #         created_at=suggestion.created_at.isoformat(),
    #         updated_at=suggestion.updated_at.isoformat(),
    #         keyword_tags=tags,
    #         confidence=confidence,
    #         candidates=[IngredientSuggestionCandidateResponse(**item) for item in candidates],
    #     )


class IngredientImportExportService:
    def __init__(self):
        self.orm = IngredientORM()

    def export_template_excel(self, user: TUser) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ingredients"

        allowed_categories = [value for value, _ in getattr(IngredientCategoryEnum, "choices")]
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(allowed_categories)}"',
            allow_blank=False,
        )
        worksheet.add_data_validation(dv)
        dv.add("B2:B1000")

        worksheet.append(IMPORT_COLUMNS)
        worksheet.append(
            [
                "Ca rot",
                "VEGETABLE",
                100,
                41,
                0.9,
                0.2,
                9.6,
                2.8,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def normalize_name(self, name: str) -> str:
        return " ".join(name.strip().lower().split())

    def import_from_excel(self, user: TUser, file: UploadedFile | None) -> IngredientImportResult:
        if not file:
            raise IngredientImportFileRequired

        file_name = file.name or ""
        if not file_name.lower().endswith(".xlsx"):
            raise IngredientImportFileInvalid

        try:
            workbook = load_workbook(filename=file, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            raise IngredientImportFileInvalid from exc

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise IngredientImportFileInvalid

        normalized_headers = [str(h).strip() if h is not None else "" for h in header_row]
        if normalized_headers[: len(IMPORT_COLUMNS)] != IMPORT_COLUMNS:
            raise IngredientImportFileInvalid

        # =========================
        # cache existing names (OPTIMIZE)
        # =========================
        existing_names = {
            self.normalize_name(name)
            for name in Ingredient.objects.filter(deleted=False)
            .values_list("name", flat=True)
        }

        created_count = 0
        failed_count = 0
        errors: list[IngredientImportError] = []
        total_rows = 0

        # =========================
        # LOOP ROWS
        # =========================
        for excel_row_index, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row_data: dict[str, Any] = {}

            for idx, column_name in enumerate(IMPORT_COLUMNS):
                row_data[column_name] = row_values[idx] if idx < len(row_values) else None

            # skip empty row
            if all(v in (None, "") for v in row_data.values()):
                continue

            total_rows += 1

            # strip string
            row_data = {
                key: (value.strip() if isinstance(value, str) else value)
                for key, value in row_data.items()
            }

            # =========================
            # VALIDATION
            # =========================
            name = row_data.get("name")
            category = row_data.get("category")

            if not name:
                failed_count += 1
                errors.append(
                    IngredientImportError(
                        row=excel_row_index,
                        message="name is required"
                    )
                )
                continue

            if not category:
                failed_count += 1
                errors.append(
                    IngredientImportError(
                        row=excel_row_index,
                        message="category is required"
                    )
                )
                continue

            name_clean = self.normalize_name(name)

            # =========================
            # DUPLICATE CHECK (NO CREATE)
            # =========================
            if name_clean in existing_names:
                failed_count += 1
                errors.append(
                    IngredientImportError(
                        row=excel_row_index,
                        message=f"Duplicate ingredient name: {name}"
                    )
                )
                continue

            # mark as existing (avoid duplicates in same file)
            existing_names.add(name_clean)

            normalized_row = {
                key: value
                for key, value in row_data.items()
                if value not in (None, "")
            }

            try:
                payload = IngredientSchema(**normalized_row)

                # =========================
                # CREATE
                # =========================
                Ingredient.objects.create(
                    **payload.dict(exclude={"name"}),
                    name=name_clean,
                    owner=user,
                    updater=user
                )

                created_count += 1

            except Exception as exc:
                failed_count += 1
                errors.append(
                    IngredientImportError(
                        row=excel_row_index,
                        message=str(exc)
                    )
                )

        return IngredientImportResult(
            total_rows=total_rows,
            created_count=created_count,
            failed_count=failed_count,
            errors=errors,
        )


class IngredientPreferenceService:
    def __init__(self, query_service: IngredientQueryService):
        self.orm = IngredientORM()
        self.query_service = query_service

    @staticmethod
    def _to_preference_response(ingredient) -> UserIngredientPreferenceResponse:
        return UserIngredientPreferenceResponse(
            ingredient_uid=ingredient.uid,
            ingredient_name=ingredient.name,
            category=ingredient.category,
        )
    @sync_user_feature("user", update_fields=["fav_ingredient"])
    def add_favourite(self, user: TUser, payload: UserIngredientPreferenceSchema):
        ingredient = self.query_service.get_by_uid(uid=payload.ingredient_uid)
        self.orm.add_favourite_ingredient(user=user, ingredient=ingredient)
        return self._to_preference_response(ingredient)

    @sync_user_feature("user", update_fields=["fav_ingredient"])
    def remove_favourite(self, user: TUser, ingredient_uid: UUID) -> bool:
        ingredient = self.query_service.get_by_uid(uid=ingredient_uid)
        return self.orm.remove_favourite_ingredient(user=user, ingredient=ingredient)

    def list_favourites(self, user: TUser):
        return [
            self._to_preference_response(item)
            for item in self.orm.list_favourite_ingredients(user=user)
        ]

    @sync_user_feature("user", update_fields=["allergy"])
    def add_allergic(self, user: TUser, payload: UserIngredientPreferenceSchema):
        ingredient = self.query_service.get_by_uid(uid=payload.ingredient_uid)
        self.orm.add_allergic_ingredient(user=user, ingredient=ingredient)
        return self._to_preference_response(ingredient)

    @sync_user_feature("user", update_fields=["allergy"])
    def remove_allergic(self, user: TUser, ingredient_uid: UUID) -> bool:
        ingredient = self.query_service.get_by_uid(uid=ingredient_uid)
        return self.orm.remove_allergic_ingredient(user=user, ingredient=ingredient)

    def list_allergic(self, user: TUser):
        return [
            self._to_preference_response(item)
            for item in self.orm.list_allergic_ingredients(user=user)
        ]


class IngredientService:
    def __init__(self):
        self.query = IngredientQueryService()
        self.command = IngredientCommandService(query_service=self.query)
        self.suggestion = IngredientSuggestionService()
        self.import_export = IngredientImportExportService()
        self.preference = IngredientPreferenceService(query_service=self.query)

    def create_new_ingredient(self, user: TUser, payload: IngredientSchema):
        return self.command.create_ingredient(user=user, payload=payload)

    def get_all_ingredients(self, filter: FilterIngredientSchema, order_by: OrderByIngredientSchema):
        return self.query.get_all(filter=filter, order_by=order_by)
    
    def get_all_ingredients_for_chef(self, user: TUser, filter: FilterIngredientSchema, order_by: OrderByIngredientSchema):
        
        return self.query.get_all_for_chef(user=user, filter=filter, order_by=order_by)

    def get_ingredient_by_uid(self, uid: UUID):
        return self.query.get_by_uid(uid=uid)

    def update_ingredient(self, user: TUser, uid: UUID, payload: IngredientSchema):
        return self.command.update_ingredient(user=user, uid=uid, payload=payload)

    def soft_delete_ingredient(self, user: TUser, uid: UUID):
        return self.command.soft_delete_ingredient(user=user, uid=uid)

    def hard_delete_ingredient(self, uid: UUID):
        return self.command.hard_delete_ingredient(uid=uid)

    def restore_ingredient(self, user: TUser, uid: UUID):
        return self.command.restore_ingredient(user=user, uid=uid)

    def search_ingredients(
        self,
        user: TUser,
        query: str,
        filter: FilterIngredientSchema,
        limit: int = 10,
    ):
        return self.query.search(
            user=user,
            query=query,
            filter=filter,
            limit=limit,
        )

    def autocomplete_ingredients(self, query: str, limit: int = 10):
        return self.query.autocomplete(query=query, limit=limit)

    def create_ingredient_alias(self, user: TUser, payload: IngredientAliasCreateSchema):
        return self.command.create_alias(user=user, payload=payload)

    def list_ingredient_aliases(self, search: str | None = None):
        return self.query.list_aliases(search=search)

    def create_ingredient_suggestion(self, user: TUser, payload: DishIngredientSuggestionSchema):
        return self.suggestion.create_suggestion(user=user, payload=payload)

    def list_ingredient_suggestions(
        self,
        filter: IngredientSuggestionFilterSchema | None = None,
        order_by: IngredientSuggestionOrderBySchema | None = None,
    ):
        status = getattr(filter, "status", None) if filter else None
        search = getattr(filter, "search", None) if filter else None
        category = getattr(filter, "category", None) if filter else None
        return self.suggestion.list_suggestions(
            status=status,
            search=search,
            category=category,
            order_by=order_by,
        )

    def get_my_ingredient_suggestions(
        self,
        user: TUser,
        filter: IngredientSuggestionFilterSchema | None = None,
        order_by: IngredientSuggestionOrderBySchema | None = None,
    ):
        status = getattr(filter, "status", None) if filter else None
        search = getattr(filter, "search", None) if filter else None
        category = getattr(filter, "category", None) if filter else None
        return self.suggestion.list_suggestions(
            status=status,
            search=search,
            category=category,
            order_by=order_by,
            created_by_id=user.id,
        )

    def approve_ingredient_suggestion(
        self,
        user: TUser,
        uid: UUID,
        payload: IngredientSuggestionResolveSchema,
    ):
        return self.suggestion.approve(user=user, uid=uid, payload=payload)

    def approve_new_ingredient_suggestion(
        self,
        user: TUser,
        uid: UUID,
        payload: IngredientSuggestionApproveNewSchema,
    ):
        return self.suggestion.approve_new(user=user, uid=uid, payload=payload)

    def approve_alias_ingredient_suggestion(
        self,
        user: TUser,
        uid: UUID,
        payload: IngredientSuggestionApproveAliasSchema,
    ):
        return self.suggestion.approve_alias(user=user, uid=uid, payload=payload)

    def reject_ingredient_suggestion(self, user: TUser, uid: UUID, rejection_reason: str):
        return self.suggestion.reject(user=user, uid=uid, rejection_reason=rejection_reason)

    def export_ingredient_template_excel(self, user: TUser) -> bytes:
        return self.import_export.export_template_excel(user=user)

    def import_ingredients_from_excel(self, user: TUser, file: UploadedFile | None) -> IngredientImportResult:
        return self.import_export.import_from_excel(user=user, file=file)

    def add_favourite_ingredient(self, user: TUser, payload: UserIngredientPreferenceSchema):
        return self.preference.add_favourite(user=user, payload=payload)

    def remove_favourite_ingredient(self, user: TUser, ingredient_uid: UUID):
        return self.preference.remove_favourite(user=user, ingredient_uid=ingredient_uid)

    def list_favourite_ingredients(self, user: TUser):
        return self.preference.list_favourites(user=user)

    def add_allergic_ingredient(self, user: TUser, payload: UserIngredientPreferenceSchema):
        return self.preference.add_allergic(user=user, payload=payload)

    def remove_allergic_ingredient(self, user: TUser, ingredient_uid: UUID):
        return self.preference.remove_allergic(user=user, ingredient_uid=ingredient_uid)

    def list_allergic_ingredients(self, user: TUser):
        return self.preference.list_allergic(user=user)
